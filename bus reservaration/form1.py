import openpyxl
import xlwt

class Bus:  # class creation

    def __init__(self): # constructor
        '''
        self.name
        self.age
        self.phno
        '''
    def reserv(self):   # function for ticket reservration
        print("Do You want to book Ticket?")  # Asking for ticket booking

        option=input("Enter Yes or No: ")

        if option == 'N' or option == "n" or option == "No" or option == "no":
            print("Thank you for visiting. ")

        elif option == 'Y' or option == 'y' or option == "Yes" or option == "yes":
            print("Welcome to System")
            print("please select Bus Type")  # for selecting bus type
            print("1. Ac_Sleeper. \n 2. Ac_ Seater. \n 3. NonAc_Sleeper. \n 4. NonAc_Seater.")

            b = int(input("Select bus type: "))
            if b == 1:
                print("Available Seats")

                #l3 = [['A1', 'A2', 'A3', 'A4'], ['B1', 'B2', 'B3', 'B4'], ['C1', 'C2', 'C3', 'C4']]
                l3= ['A1', 'A2', 'A3', 'A4', 'B1', 'B2', 'B3', 'B4', 'C1', 'C2', 'C3', 'C4']
                print(l3)

                book = openpyxl.Workbook()  # creating Excel sheet
                ac_sleeper = book.active

                row=0
                column=0
                for item in l3:
                    ac_sleeper.write(row, column, item)
                    row+=1
'''
                for i in range(1, (len(l3) + 1)):
                    for j in range(1, (len(l3) + 2)):
                        ac_sleeper.cell(row=i, column=j).value = l3   #[i-1][j-1]
                book.save("Ac-sleeper.xlsx")

                book = openpyxl.load_workbook("Ac-sleeper.xlsx")  # to read excel file
                ac_sleeper = book.active
                s = input("please select seat number: ")

                for ele in ac_sleeper:
                    if ele.value == s:
                        print("Seat successfully booked.")
                    else:
                        print("Seats are not avialable")

                book.save("Ac-sleeper.xlsx")


                elif b == 2:
                    print("Available Seats")

                    l2 = [[1, 2, 3, 4, 5, 6], [6, 7, 8, 9, 10, 11], [11, 12, 13, 14, 15, 16], [16, 17, 18, 19, 20, 21],
                          [21, 22, 23, 24, 25, 26]]

                    book = openpyxl.Workbook()  # creating Excel sheet
                    ac_seater = book.active

                    for i in range(1, (len(l2) + 1)):
                        # for j in range(1,(len(l2)+2)):
                        for j in range(1, len(l2[i - 1])):
                            ac_seater.cell(row=i, column=j).value = l2[i - 1][j - 1]
                    book.save("Ac-seater.xlsx")

                elif b==3:
                     print("Available Seats")

                     l3 = [['A1', 'A2', 'A3', 'A4'], ['B1', 'B2', 'B3', 'B4'], ['C1', 'C2', 'C3', 'C4']]
                     print(l3)

                     book = openpyxl.Workbook()  # creating Excel sheet
                     nonac_sleeper = book.active

                     for i in range(1, (len(l3) + 1)):
                         for j in range(1, (len(l3) + 2)):
                             nonac_sleeper.cell(row=i, column=j).value = l3[i - 1][j - 1]
                     book.save("NonAc-sleeper.xlsx")

                elif b==4:
                    print("Available Seats")

                    l4 = [[1, 2, 3, 4, 5, 6], [6, 7, 8, 9, 10, 11], [11, 12, 13, 14, 15, 16], [16, 17, 18, 19, 20, 21],
                          [21, 22, 23, 24, 25, 26]]

                    book = openpyxl.Workbook()  # creating Excel sheet
                    nonac_seater = book.active

                    for i in range(1, (len(l4) + 1)):
                        for j in range(1, len(l4[i - 1])):
                            nonac_seater.cell(row=i, column=j).value = l4[i - 1][j - 1]
                    book.save("NonAc-seater.xlsx")
                else:
                    print("Seats are not avialable.")


    else:
        print("Not Recognised.")




'''


b=Bus()  # Class object Creation
b.reserv() # calling function using class object